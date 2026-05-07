# -*- coding: utf-8 -*-
"""
Submission API routes.

Endpoints:
  POST   /submission/upload             – Upload submission PDF for evaluation
  GET    /submission/{id}/status        – Check evaluation progress
  GET    /submission/{id}/report        – Get full evaluation report
  GET    /tender/{tender_id}/submissions – List all submissions for a tender
"""

from __future__ import annotations

import logging
import re
import json
from datetime import datetime
from io import BytesIO
from math import ceil
from typing import Any

from fastapi import APIRouter, File, Form, UploadFile, HTTPException, Query
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook

from app.models.db import get_db

_DEVANAGARI_RE = re.compile(r"[\u0900-\u097F]")
_GENERIC_HEADING_RE = re.compile(r"^Table_\d+$", re.IGNORECASE)

router = APIRouter(tags=["submission"])
logger = logging.getLogger(__name__)

_DETAIL_TABS = [
    {"tab_key": "technical_infrastructure", "tab_label": "Technical Infrastructure", "is_default": True},
    {"tab_key": "security_compliance", "tab_label": "Security & Compliance", "is_default": False},
    {"tab_key": "operational_continuity", "tab_label": "Operational Continuity", "is_default": False},
]

_TAB_TO_CATEGORIES = {
    "technical_infrastructure": {"Technical Specifications", "Experience & Capability"},
    "security_compliance": {"Legal & Compliance"},
    "operational_continuity": {"Commercial / Tender Terms", "Financial Thresholds & Stability"},
}


def _to_percent(value) -> float | None:
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v <= 1.0:
        v *= 100.0
    return round(v, 1)


def _normalize_submission_verdict(verdict: str | None) -> str | None:
    if not verdict:
        return None
    mapping = {
        "ELIGIBLE": "QUALIFIED",
        "NOT_ELIGIBLE": "DISQUALIFIED",
        "MANUAL_REVIEW": "MANUAL_REVIEW",
    }
    return mapping.get(verdict, verdict)


def _status_match(sub_status: str | None, status_filter: str) -> bool:
    status = (sub_status or "").upper()
    f = status_filter.upper()
    if f == "ALL":
        return True
    if f == "READY":
        return status == "READY"
    if f == "FAILED":
        return status == "FAILED"
    if f == "PROCESSING":
        return status not in {"READY", "FAILED"}
    return False


def _safe_iso(ts) -> str:
    if isinstance(ts, datetime):
        return ts.isoformat()
    return str(ts or "")


def _human_bidder_name(file_name: str | None) -> str | None:
    if not file_name:
        return None
    base = file_name.rsplit(".", 1)[0]
    return " ".join(base.replace("_", " ").replace("-", " ").split()) or file_name


def _build_submission_report_xlsx(report: dict[str, Any]) -> bytes:
    """
    Build a real XLSX file for a single submission report.
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Field", "Value"])
    summary = report.get("summary", {}) or {}
    ws_summary.append(["submission_id", report.get("submission_id")])
    ws_summary.append(["tender_id", report.get("tender_id")])
    for k in (
        "evaluation_timestamp",
        "overall_score",
        "verdict",
        "total_criteria_evaluated",
        "mandatory_failures_count",
    ):
        ws_summary.append([k, summary.get(k)])

    # Sheet 2: Category Scores
    ws_cat = wb.create_sheet("Category Scores")
    ws_cat.append(
        [
            "category",
            "deterministic_score",
            "llm_score",
            "final_score",
            "total_criteria",
            "passed",
            "failed",
            "needs_review",
        ]
    )
    for category, payload in (report.get("category_scores") or {}).items():
        payload = payload or {}
        ws_cat.append(
            [
                category,
                payload.get("deterministic_score"),
                payload.get("llm_score"),
                payload.get("final_score"),
                payload.get("total_criteria"),
                payload.get("passed"),
                payload.get("failed"),
                payload.get("needs_review"),
            ]
        )

    # Sheet 3: Criteria Details
    ws_criteria = wb.create_sheet("Criteria")
    ws_criteria.append(
        [
            "criterion_evidence_id",
            "category",
            "sub_component",
            "expected_value",
            "found_value",
            "deterministic_score",
            "verdict",
            "reasoning",
            "matched_evidence_ids",
            "matched_pages",
        ]
    )
    for item in (report.get("criteria_details") or []):
        item = item or {}
        ws_criteria.append(
            [
                item.get("criterion_evidence_id"),
                item.get("category"),
                item.get("sub_component"),
                item.get("expected_value"),
                item.get("found_value"),
                item.get("deterministic_score"),
                item.get("verdict"),
                item.get("reasoning"),
                ", ".join(str(x) for x in (item.get("matched_evidence_ids") or [])),
                ", ".join(str(x) for x in (item.get("matched_pages") or [])),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_consolidated_report_xlsx(consolidated: dict[str, Any]) -> bytes:
    """
    Build a real XLSX file for consolidated tender report.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(
        [
            "submission_id",
            "bidder_name",
            "bidder_file_name",
            "status",
            "upload_timestamp",
            "overall_score",
            "verdict",
            "mandatory_failures",
            "technical_specifications",
            "financial_thresholds_stability",
            "experience_capability",
            "legal_compliance",
            "commercial_tender_terms",
        ]
    )

    for row in (consolidated.get("items") or []):
        row = row or {}
        cat = row.get("category_scores") or {}
        ws.append(
            [
                row.get("submission_id"),
                row.get("bidder_name"),
                row.get("bidder_file_name"),
                row.get("status"),
                row.get("upload_timestamp"),
                row.get("overall_score"),
                row.get("verdict"),
                row.get("mandatory_failures"),
                cat.get("Technical Specifications"),
                cat.get("Financial Thresholds & Stability"),
                cat.get("Experience & Capability"),
                cat.get("Legal & Compliance"),
                cat.get("Commercial / Tender Terms"),
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_canonical_ev_map(tender_id: str) -> dict[str, Any]:
    """
    Return a flat dict {evidence_id: canonical_item} for all items stored
    inside the canonical_tenders document for this tender.
    """
    db = get_db()
    doc = db.canonical_tenders.find_one({"tender_id": tender_id}, {"_id": 0}) or {}
    ev_map: dict[str, Any] = {}
    _IGNORE = {"tender_id", "total_classified", "total_ignored", "batch_count",
               "created_at", "pdf_path", "fields", "evidence", "Ignore"}
    for key, bucket in doc.items():
        if key in _IGNORE or not isinstance(bucket, list):
            continue
        for item in bucket:
            eid = item.get("evidence_id")
            if eid:
                ev_map[eid] = item
    return ev_map


def _derive_item_title_and_bid_text(
    ev: dict[str, Any] | None,
) -> tuple[str, str | None]:
    """
    Same rule as tender details:
      table rows  → item_title = key_norm (fallback: heading), bid_text = text_norm
      other rows  → item_title = heading,                      bid_text = text_norm
    Hindi-only headings are replaced with sub_component fallback.
    """
    if not ev:
        return "Criterion", None

    source = (ev.get("source") or "").strip().lower()
    heading = (ev.get("heading") or "").strip()
    key = (ev.get("key_norm") or "").strip()
    text = (ev.get("text_norm") or ev.get("text_raw") or "").strip()
    sub = (ev.get("sub_component") or "").strip()

    def _clean_heading(h: str) -> str | None:
        if not h:
            return None
        if _DEVANAGARI_RE.search(h) or _GENERIC_HEADING_RE.match(h):
            return None
        return h

    if source == "table":
        name = key or _clean_heading(heading) or sub or "Criterion"
    else:
        name = _clean_heading(heading) or sub or "Criterion"

    return name[:120], (text or None)


@router.post("/submission/upload")
async def upload_submission(
    file: UploadFile = File(...),
    tender_id: str = Form(...),
):
    """Upload a bidder submission PDF and trigger full evaluation."""
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    # Check that the tender exists
    db = get_db()
    tender = db.tenders.find_one({"_id": tender_id})
    if not tender:
        raise HTTPException(status_code=404, detail=f"Tender {tender_id} not found.")

    # Check that canonical embeddings exist
    emb_count = db.embeddings.count_documents({"tender_id": tender_id})
    if emb_count == 0:
        raise HTTPException(
            status_code=400,
            detail=f"Tender {tender_id} has no canonical embeddings. "
                   "Upload and process the tender PDF first.",
        )

    from app.services.submission_service import evaluate_submission

    try:
        result = await evaluate_submission(tender_id, file)
        return JSONResponse(content=result)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/submission/{submission_id}/status")
async def get_submission_status(submission_id: str):
    """Check the evaluation status of a submission."""
    db = get_db()
    submission = db.submissions.find_one(
        {"_id": submission_id}, {"_id": 1, "status": 1, "tender_id": 1}
    )
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    return {
        "submission_id": submission_id,
        "tender_id": submission.get("tender_id"),
        "status": submission.get("status"),
    }


@router.get("/submission/{submission_id}/report")
async def get_submission_report(submission_id: str):
    """Get the full evaluation report for a submission."""
    db = get_db()
    report = db.evaluation_reports.find_one({"submission_id": submission_id})

    if not report:
        raise HTTPException(
            status_code=404,
            detail="Evaluation report not found. Check submission status.",
        )

    # Remove MongoDB internal _id for JSON serialization
    report.pop("_id", None)
    return JSONResponse(content=report)


@router.get("/submission/{submission_id}/report/download")
async def download_submission_report(
    submission_id: str,
    format: str = Query("pdf", description="pdf | xlsx | json"),
):
    """
    Download a single submission report in the requested format.
    Current implementation exports the same report payload with different
    filenames/content-types for frontend download workflows.
    """
    fmt = (format or "pdf").lower()
    if fmt not in {"pdf", "xlsx", "json"}:
        raise HTTPException(status_code=400, detail="Invalid format. Allowed: pdf, xlsx, json.")

    db = get_db()
    submission = db.submissions.find_one({"_id": submission_id}, {"_id": 1, "status": 1})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")
    if (submission.get("status") or "").upper() != "READY":
        raise HTTPException(status_code=409, detail="Report not ready.")

    report = db.evaluation_reports.find_one({"submission_id": submission_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found.")

    if fmt == "json":
        payload = json.dumps(report, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        media_type = "application/json"
        filename = f"submission_{submission_id}_report.json"
    elif fmt == "xlsx":
        payload = _build_submission_report_xlsx(report)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"submission_{submission_id}_report.xlsx"
    else:
        payload = json.dumps(report, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        media_type = "application/pdf"
        filename = f"submission_{submission_id}_report.pdf"

    return Response(
        content=payload,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tender/{tender_id}/submissions")
async def list_submissions(
    tender_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    status: str = Query("READY", description="READY | PROCESSING | FAILED | ALL"),
    search: str = Query(""),
    sort_by: str = Query("overall_score", description="overall_score | upload_timestamp | bidder_name"),
    sort_order: str = Query("desc", description="asc | desc"),
):
    """List submissions for a tender with filtering, sorting and pagination."""
    db = get_db()
    if not db.tenders.find_one({"_id": tender_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail=f"Tender {tender_id} not found.")

    status_u = status.upper()
    if status_u not in {"READY", "PROCESSING", "FAILED", "ALL"}:
        raise HTTPException(status_code=422, detail="Invalid status filter.")
    if sort_by not in {"overall_score", "upload_timestamp", "bidder_name"}:
        raise HTTPException(status_code=422, detail="Invalid sort_by value.")
    order = sort_order.lower()
    if order not in {"asc", "desc"}:
        raise HTTPException(status_code=422, detail="Invalid sort_order value.")

    needle = search.strip().lower()
    submissions = list(
        db.submissions.find(
            {"tender_id": tender_id},
            {"_id": 1, "bidder_name": 1, "status": 1, "upload_timestamp": 1, "file_path": 1},
        )
    )

    # Enrich with scores from evaluation reports
    results = []
    for sub in submissions:
        sub_id = sub["_id"]
        report = db.evaluation_reports.find_one({"submission_id": sub_id}, {"_id": 0})
        summary = (report or {}).get("summary", {})
        category_scores_raw = (report or {}).get("category_scores", {})
        category_scores = {
            cat: _to_percent((score_obj or {}).get("final_score"))
            for cat, score_obj in category_scores_raw.items()
        }
        file_name = sub.get("bidder_name")

        entry = {
            "submission_id": sub_id,
            "bidder_name": _human_bidder_name(file_name),
            "bidder_file_name": file_name,
            "status": sub.get("status"),
            "upload_timestamp": _safe_iso(sub.get("upload_timestamp")),
            "overall_score": _to_percent(summary.get("overall_score")),
            "verdict": _normalize_submission_verdict(summary.get("verdict")),
            "category_scores": category_scores,
            "mandatory_failures": len((report or {}).get("mandatory_failures", [])),
        }
        results.append(entry)

    # Filters
    results = [r for r in results if _status_match(r.get("status"), status_u)]
    if needle:
        results = [
            r
            for r in results
            if needle in (r.get("bidder_name") or "").lower()
            or needle in (r.get("bidder_file_name") or "").lower()
            or needle in (r.get("submission_id") or "").lower()
        ]

    # Sort
    reverse = order == "desc"
    if sort_by == "overall_score":
        results.sort(key=lambda x: x.get("overall_score") or 0, reverse=reverse)
    elif sort_by == "upload_timestamp":
        results.sort(key=lambda x: x.get("upload_timestamp") or "", reverse=reverse)
    else:
        results.sort(key=lambda x: (x.get("bidder_name") or "").lower(), reverse=reverse)

    total = len(results)
    start = (page - 1) * page_size
    items = results[start : start + page_size]

    return {
        "tender_id": tender_id,
        "total_submissions": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


@router.get("/tender/{tender_id}/comparison/summary")
async def comparison_summary(tender_id: str):
    db = get_db()
    if not db.tenders.find_one({"_id": tender_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail=f"Tender {tender_id} not found.")

    subs = list(db.submissions.find({"tender_id": tender_id}, {"_id": 1}))
    reports = []
    for s in subs:
        rep = db.evaluation_reports.find_one({"submission_id": s["_id"]}, {"summary.overall_score": 1})
        if rep and rep.get("summary", {}).get("overall_score") is not None:
            reports.append(rep)

    total = len(subs)
    score_values = [_to_percent(r["summary"]["overall_score"]) for r in reports]
    score_values = [x for x in score_values if x is not None]
    avg_match = round(sum(score_values) / len(score_values), 1) if score_values else 0.0

    if avg_match >= 85:
        critical_path = "LOW"
    elif avg_match >= 65:
        critical_path = "MEDIUM"
    else:
        critical_path = "HIGH"

    stability_score = round(max(0.0, 100.0 - (abs(80.0 - avg_match) * 0.6)), 1)
    insight = (
        f"Evaluated {total} vendors for this tender. "
        f"Average match is {avg_match}%, with overall risk level {critical_path}."
    )

    return {
        "tender_id": tender_id,
        "total_submissions": total,
        "average_match_percent": avg_match,
        "critical_path": critical_path,
        "stability_score": stability_score,
        "insight_text": insight,
    }


@router.get("/tender/{tender_id}/report/download")
async def download_tender_consolidated_report(
    tender_id: str,
    format: str = Query("xlsx", description="xlsx | json"),
):
    """
    Consolidated tender report download across all submissions for a tender.
    """
    fmt = (format or "xlsx").lower()
    if fmt not in {"xlsx", "json"}:
        raise HTTPException(status_code=400, detail="Invalid format. Allowed: xlsx, json.")

    db = get_db()
    if not db.tenders.find_one({"_id": tender_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail=f"Tender {tender_id} not found.")

    submissions = list(
        db.submissions.find(
            {"tender_id": tender_id},
            {"_id": 1, "bidder_name": 1, "status": 1, "upload_timestamp": 1},
        )
    )
    if not submissions:
        raise HTTPException(status_code=404, detail="No submissions found for this tender.")

    consolidated_items = []
    for sub in submissions:
        sub_id = sub["_id"]
        report = db.evaluation_reports.find_one(
            {"submission_id": sub_id},
            {"_id": 0, "summary": 1, "mandatory_failures": 1, "category_scores": 1},
        ) or {}
        summary = report.get("summary", {})
        consolidated_items.append(
            {
                "submission_id": sub_id,
                "bidder_name": _human_bidder_name(sub.get("bidder_name")),
                "bidder_file_name": sub.get("bidder_name"),
                "status": sub.get("status"),
                "upload_timestamp": _safe_iso(sub.get("upload_timestamp")),
                "overall_score": _to_percent(summary.get("overall_score")),
                "verdict": _normalize_submission_verdict(summary.get("verdict")),
                "mandatory_failures": len(report.get("mandatory_failures", [])),
                "category_scores": {
                    cat: _to_percent((score_obj or {}).get("final_score"))
                    for cat, score_obj in (report.get("category_scores", {}) or {}).items()
                },
            }
        )

    consolidated = {
        "tender_id": tender_id,
        "total_submissions": len(consolidated_items),
        "items": consolidated_items,
    }
    if fmt == "json":
        payload = json.dumps(consolidated, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        media_type = "application/json"
        filename = f"tender_{tender_id}_consolidated_report.json"
    else:
        payload = _build_consolidated_report_xlsx(consolidated)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"tender_{tender_id}_consolidated_report.xlsx"

    return Response(
        content=payload,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tender/{tender_id}/comparison/filters")
async def comparison_filters(tender_id: str):
    db = get_db()
    if not db.tenders.find_one({"_id": tender_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail=f"Tender {tender_id} not found.")
    return {
        "vendor_context_options": [
            {"value": "ALL", "label": "All Selected Vendors"},
            {"value": "QUALIFIED_ONLY", "label": "Qualified Only"},
            {"value": "FAILED_ONLY", "label": "Disqualified Only"},
        ],
        "risk_levels": ["LOW", "MEDIUM", "HIGH"],
    }


@router.get("/submission/{submission_id}/details/meta")
async def submission_details_meta(submission_id: str):
    db = get_db()
    submission = db.submissions.find_one({"_id": submission_id})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    tender_id = submission.get("tender_id")
    tender = db.tenders.find_one({"_id": tender_id}) or {}
    report = db.evaluation_reports.find_one({"submission_id": submission_id}, {"_id": 0}) or {}
    summary = report.get("summary", {})
    canonical = db.canonical_tenders.find_one({"tender_id": tender_id}, {"_id": 0}) or {}
    fields = canonical.get("fields") or {}

    return {
        "submission": {
            "submission_id": submission_id,
            "tender_id": tender_id,
            "tender_reference": tender.get("reference") or tender_id,
            "tender_name": tender.get("name") or tender_id,
            "bidder_name": _human_bidder_name(submission.get("bidder_name")),
            "overall_match_percent": _to_percent(summary.get("overall_score")) or 0.0,
            "status": submission.get("status"),
            "verdict": _normalize_submission_verdict(summary.get("verdict")),
            "category": None,
            "location": fields.get("buyer.location") or fields.get("delivery.location"),
            "due_date": (fields.get("bid.end_date_time") or "").split("T", 1)[0] or None,
        },
        "tabs": _DETAIL_TABS,
    }


@router.get("/submission/{submission_id}/details/items")
async def submission_details_items(
    submission_id: str,
    tab_key: str = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    search: str = Query(""),
    min_match: int = Query(0, ge=0, le=100),
):
    if tab_key not in _TAB_TO_CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid tab_key.")

    db = get_db()
    submission = db.submissions.find_one({"_id": submission_id}, {"_id": 1})
    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    report = db.evaluation_reports.find_one({"submission_id": submission_id}, {"_id": 0}) or {}
    tender_id = db.submissions.find_one({"_id": submission_id}, {"tender_id": 1}).get("tender_id")
    ev_map = _build_canonical_ev_map(tender_id) if tender_id else {}

    details = report.get("criteria_details", [])
    allowed_cats = _TAB_TO_CATEGORIES[tab_key]
    needle = search.strip().lower()

    filtered = []
    for idx, d in enumerate(details, 1):
        category = d.get("category")
        if category not in allowed_cats:
            continue
        match_percent = _to_percent(d.get("deterministic_score")) or 0.0
        if match_percent < min_match:
            continue

        # Look up the original canonical evidence item for this criterion
        ev = ev_map.get(d.get("criterion_evidence_id") or "")
        title, bid_mention = _derive_item_title_and_bid_text(ev)

        # Skip rows whose title is in Hindi (same filter as tender details)
        if _DEVANAGARI_RE.search(title):
            continue

        # Fallback bid_mention to stored expected_value when canonical lookup fails
        if not bid_mention:
            bid_mention = d.get("expected_value")

        vendor_evidence = d.get("found_value") or d.get("found_raw")
        verdict = (d.get("verdict") or "").upper()
        if verdict == "PASS":
            match_status = "MATCHED"
        elif verdict == "FAIL":
            match_status = "MISSING"
        else:
            match_status = "PARTIAL"

        row = {
            "item_id": f"SPEC-{idx:03d}",
            "item_title": title,
            "item_subtitle": None,
            "bid_mention_text": bid_mention,
            "vendor_evidence_text": vendor_evidence or None,
            "match_percent": int(round(match_percent)),
            "match_status": match_status,
            "is_mandatory": bool(
                d.get("verdict") == "FAIL"
                or (d.get("reasoning") and "mandatory" in str(d.get("reasoning")).lower())
            ),
            "failure_reason": d.get("reasoning") if match_status == "MISSING" else None,
        }

        if needle:
            haystack = " ".join(
                [
                    str(row.get("item_title") or ""),
                    str(row.get("bid_mention_text") or ""),
                    str(row.get("vendor_evidence_text") or ""),
                ]
            ).lower()
            if needle not in haystack:
                continue

        filtered.append(row)

    total = len(filtered)
    total_pages = ceil(total / page_size) if total else 0
    start = (page - 1) * page_size
    items = filtered[start : start + page_size]

    return {
        "submission_id": submission_id,
        "tab_key": tab_key,
        "columns": [
            {"key": "item_id", "label": "Item"},
            {"key": "bid_mention", "label": "Mention in Bid"},
            {"key": "vendor_evidence", "label": "Provided by Vendor"},
            {"key": "match_percent", "label": "Matching"},
        ],
        "items": items,
        "pagination": {
            "total_items": total,
            "total_pages": total_pages,
            "page": page,
            "page_size": page_size,
        },
    }
